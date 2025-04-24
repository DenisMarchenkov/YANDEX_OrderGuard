import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.worksheet import Worksheet
from logging_config import *


def set_common_properties(sheet: Worksheet, file_name: str, sfx, start_row=2):
    cm = 1 / 2.54
    sheet.page_margins = PageMargins(left=cm * 0.8, right=cm * 0.8, top=cm * 0.8, bottom=cm * 1.8)
    sheet.oddFooter.left.text = os.path.basename(file_name).split('.')[0] + " - " + sfx

    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(border_style='thin', color='FF000000'),
        right=Side(border_style='thin', color='FF000000'),
        top=Side(border_style='thin', color='FF000000'),
        bottom=Side(border_style='thin', color='FF000000'),
    )

    for row in range(start_row, sheet.max_row + 1):
        for column in range(1, sheet.max_column + 1):
            cell = sheet.cell(row, column)
            cell.alignment = alignment
            cell.border = thin_border
            cell.font = Font(name='Calibri', size=14)


def format_recap_sheet(sheet: Worksheet, file_name: str, sfx):
    sheet.insert_rows(1)
    sheet.cell(1, 1).font = Font(name='Calibri', size=20)
    sheet.cell(1, 1).value = os.path.basename(file_name).split('.')[0] + " - " + sfx
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 5
    sheet.print_title_rows = "2:2"


def format_merge_table_sheet(sheet: Worksheet, file_name: str, sfx):
    sheet.insert_rows(idx=1)
    sheet.cell(1, 1).font = Font(name='Calibri', size=20)
    sheet.cell(1, 1).value = f"СВОДНАЯ ТАБЛИЦА К {os.path.basename(file_name).split('.')[0]} - {sfx}"
    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 17
    sheet.column_dimensions['C'].width = 44
    sheet.column_dimensions['D'].width = 14
    sheet.column_dimensions['E'].width = 5
    sheet.print_title_rows = "2:2"

    grey_fill = PatternFill(fill_type='solid', fgColor='FF808080')
    white_font = Font(color='ffffff', name='Calibri', size=14)
    thin_border = Border(bottom=Side(border_style='thin', color='ffffff'), right=Side(border_style='thin', color='ffffff'))
    alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    date_format = 'DD.MM.YYYY'

    for row in range(3, sheet.max_row + 1):
        sheet.cell(row, 3).alignment = alignment
        sheet.cell(row, 4).number_format = date_format

        if sheet.cell(row, 5).value > 1:
            sheet.cell(row, 5).fill = PatternFill(fill_type='solid', fgColor='FF000000')
            sheet.cell(row, 5).font = white_font
            sheet.cell(row, 5).border = Border(bottom=Side(border_style='thin', color='ffffff'))

        current_value = sheet.cell(row, 2).value
        if row < sheet.max_row and current_value == sheet.cell(row + 1, 2).value:
            for col in range(2, 5):
                for r in [row, row + 1]:
                    sheet.cell(r, col).fill = grey_fill
                    sheet.cell(r, col).font = white_font
                    sheet.cell(r, col).border = thin_border


def format_orders_sheet(sheet: Worksheet):
    #Поля
    cm = 1 / 2.54
    sheet.page_margins = PageMargins(left=cm * 0.8, right=cm * 0.8, top=cm * 0.8, bottom=cm * 1.8)

    # Стили
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )
    header_align = Alignment(horizontal='center', vertical='center')
    header_font = Font(name='Calibri', size=16)

    # Вставка строк и заполнение заголовков
    sheet.insert_rows(1, 6)
    headers = ["НОМЕР ЗАКАЗА", "ДАТА ЗАКАЗА", "ДАТА ОТГРУЗКИ", "СТРОК В ЗАКАЗЕ", "ШТУК В ЗАКАЗЕ"]
    for i, text in enumerate(headers, start=1):
        sheet.cell(i, 2).value = text

    # Значения для шапки
    sheet.cell(1, 4).value = sheet.cell(8, 1).value
    sheet.cell(2, 4).value = sheet.cell(8, 8).value
    sheet.cell(2, 4).number_format = 'DD.MM.YYYY'
    sheet.cell(3, 4).value = sheet.cell(8, 9).value
    sheet.cell(3, 4).number_format = 'DD.MM.YYYY'

    # Удаление ненужных колонок
    sheet.delete_cols(1)
    sheet.delete_cols(7, 2)

    # Ширина колонок
    column_widths = [14, 16, 36, 4, 13, 14]
    for i, width in enumerate(column_widths, start=1):
        col_letter = chr(64 + i)
        sheet.column_dimensions[col_letter].width = width

    # Оформление заголовков
    for row in range(1, 6):
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        sheet.row_dimensions[row].height = 25
        for col in range(1, 4):
            cell = sheet.cell(row=row, column=col)
            cell.alignment = header_align
            cell.border = thin_border
            cell.font = header_font

    # Форматирование основной таблицы
    stat_order = []
    alignments = {
        1: Alignment(horizontal='left', vertical='center', wrap_text=True),
        2: Alignment(horizontal='left', vertical='center', wrap_text=True),
        3: Alignment(horizontal='left', vertical='center', wrap_text=True),
        4: Alignment(horizontal='center', vertical='center', wrap_text=True),
        5: Alignment(horizontal='right', vertical='center', wrap_text=True),
        6: Alignment(horizontal='right', vertical='center', wrap_text=True),
    }

    highlight_fill = PatternFill(fill_type='solid', fgColor='000000')
    highlight_font = Font(color='FFFFFF', name='Calibri', size=14)
    highlight_border = Border(bottom=Side(style='thin', color='FFFFFF'))

    for row in range(8, sheet.max_row + 1):
        for col in range(1, 7):
            cell = sheet.cell(row, col)
            cell.alignment = alignments[col]
            cell.border = thin_border
            if col == 6:
                cell.number_format = 'DD.MM.YYYY'

        val = sheet.cell(row, 4).value
        stat_order.append(val)
        if val and val > 1:
            cell = sheet.cell(row, 4)
            cell.fill = highlight_fill
            cell.font = highlight_font
            cell.border = highlight_border

    # Итоги
    sheet.cell(4, 3).value = len(stat_order)
    sheet.cell(5, 3).value = sum(stat_order)

    # Печать и футер
    sheet.oddFooter.right.text = f'Заказ {sheet.cell(1, 3).value}'
    sheet.print_options.horizontalCentered = True
    sheet.print_title_rows = "7:7"


def format_assembly_sheet(sheet: Worksheet, file_name: str, sfx):
    sheet.insert_rows(1)
    sheet.cell(1, 1).font = Font(name='Calibri', size=20)
    sheet.cell(1, 1).value = f"ЛИСТ ПОДБОРА К {os.path.basename(file_name).split('.')[0]} - {sfx}"
    sheet.page_setup.orientation = 'landscape'
    sheet.column_dimensions['A'].width = 25
    sheet.column_dimensions['B'].width = 18
    sheet.column_dimensions['C'].width = 18
    sheet.column_dimensions['D'].width = 67
    sheet.column_dimensions['E'].width = 4
    sheet.print_title_rows = "2:2"

    alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    grey_fill = PatternFill(fill_type='solid', fgColor='FF808080')
    white_font = Font(color='ffffff', name='Calibri', size=14)
    thin_border = Border(bottom=Side(border_style='thin', color='ffffff'), right=Side(border_style='thin', color='ffffff'))

    for row in range(3, sheet.max_row + 1):
        sheet.cell(row, 4).alignment = alignment
        if sheet.cell(row, 5).value > 1:
            sheet.cell(row, 5).fill = PatternFill(fill_type='solid', fgColor='FF000000')
            sheet.cell(row, 5).font = white_font
            sheet.cell(row, 5).border = Border(bottom=Side(border_style='thin', color='ffffff'))

        current_value = sheet.cell(row, 1).value
        if row < sheet.max_row and current_value == sheet.cell(row + 1, 1).value:
            for col in range(1, 5):
                for r in [row, row + 1]:
                    sheet.cell(r, col).fill = grey_fill
                    sheet.cell(r, col).font = white_font
                    sheet.cell(r, col).border = thin_border


def format_report_sheets(report_path: str, suffix_shop: str):
    wb = openpyxl.load_workbook(report_path)

    formatters = [
        format_recap_sheet,
        format_merge_table_sheet,
        format_assembly_sheet,
    ]

    for i, sheet_name in enumerate(wb.sheetnames):
        sheet = wb[sheet_name]

        if i < 3:
            # Первые три листа: общие свойства + своя функция форматирования
            set_common_properties(sheet, report_path, suffix_shop)
            formatters[i](sheet, report_path, suffix_shop)
        else:
            # Остальные: заказы
            format_orders_sheet(sheet)

    wb.save(report_path)